import fitz
from docx import Document as DocxDocument
import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List
from loguru import logger
from PIL import Image
import pytesseract
import io

# Set Tesseract path (adjust if needed)
import os
if os.path.exists(r'C:\Program Files\Tesseract-OCR\tesseract.exe'):
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
elif os.path.exists(r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'):
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'

class ContentExtractor:
    def __init__(self):
        self.logger = logger
        self.supported_types = {
            '.pdf': self.extract_from_pdf,
            '.docx': self.extract_from_docx,
            '.xlsx': self.extract_from_excel,
            '.csv': self.extract_from_csv,
            '.png': self.extract_from_image,
            '.jpg': self.extract_from_image,
            '.jpeg': self.extract_from_image,
            '.tiff': self.extract_from_image,
            '.tif': self.extract_from_image,
        }
    
    async def extract(self, file_path: Path) -> Dict:
        """Extract content from document"""
        ext = file_path.suffix.lower()
        
        if ext not in self.supported_types:
            self.logger.warning(f"Unsupported file type: {ext}")
            return {'text': f'[Unsupported type: {ext}]', 'error': f'Unsupported: {ext}', 'page_count': 0, 'is_scanned': False}
        
        try:
            extractor = self.supported_types[ext]
            result = await extractor(file_path)
            result['file_type'] = ext
            result['file_size'] = file_path.stat().st_size
            return result
        except Exception as e:
            self.logger.error(f"Extraction failed for {ext}: {e}")
            return {'text': f'[Error: {str(e)}]', 'error': str(e), 'page_count': 0, 'is_scanned': False}
    
    async def extract_from_image(self, file_path: Path) -> Dict:
        """Extract text from image using OCR"""
        self.logger.info(f"Running OCR on image: {file_path.name}")
        
        try:
            # Open image
            img = Image.open(str(file_path))
            
            # Convert to RGB if needed
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Perform OCR
            text = pytesseract.image_to_string(img)
            
            # Get confidence
            data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
            confidences = [int(c) for c in data['conf'] if c != '-1']
            avg_confidence = sum(confidences) / len(confidences) / 100 if confidences else 0
            
            word_count = len(text.split())
            
            self.logger.info(f"OCR complete: {word_count} words, confidence: {avg_confidence:.2%}")
            
            if word_count < 5:
                self.logger.warning("Very little text detected in image")
            
            return {
                'text': text,
                'pages': [{'page': 1, 'text': text, 'word_count': word_count}],
                'page_count': 1,
                'is_scanned': True,
                'ocr_confidence': avg_confidence,
                'metadata': {'ocr_confidence': avg_confidence}
            }
            
        except Exception as e:
            self.logger.error(f"OCR failed: {e}")
            self.logger.error("Make sure Tesseract OCR is installed: https://github.com/UB-Mannheim/tesseract/wiki")
            return {
                'text': f'[OCR Error: Tesseract not found. Install from https://github.com/UB-Mannheim/tesseract/wiki]',
                'pages': [],
                'page_count': 0,
                'is_scanned': True,
                'metadata': {}
            }
    
    async def extract_from_pdf(self, file_path: Path) -> Dict:
        """Extract content from PDF"""
        self.logger.info(f"Extracting PDF: {file_path.name}")
        
        full_text = []
        pages_content = []
        page_count = 0
        is_scanned = True
        
        try:
            doc = fitz.open(str(file_path))
            page_count = len(doc)
            
            for page_num in range(page_count):
                try:
                    page = doc[page_num]
                    text = page.get_text()
                    
                    # If little text, try OCR on the page
                    if len(text.strip()) < 50:
                        self.logger.info(f"Page {page_num+1} appears scanned, running OCR...")
                        pix = page.get_pixmap(dpi=200)
                        img = Image.open(io.BytesIO(pix.tobytes("png")))
                        text = pytesseract.image_to_string(img)
                    
                    if text and len(text.strip()) > 0:
                        full_text.append(text)
                        pages_content.append({
                            'page': page_num + 1,
                            'text': text,
                            'word_count': len(text.split())
                        })
                except Exception as e:
                    self.logger.warning(f"Error on page {page_num+1}: {e}")
            
            total_text = '\n'.join(full_text)
            is_scanned = len(total_text.strip()) < 50
            
            doc.close()
            
            return {
                'text': total_text,
                'pages': pages_content,
                'page_count': page_count,
                'is_scanned': is_scanned,
                'metadata': {}
            }
            
        except Exception as e:
            self.logger.error(f"PDF error: {e}")
            try:
                if 'doc' in locals():
                    doc.close()
            except:
                pass
            return {'text': f'[Error: {e}]', 'pages': [], 'page_count': 0, 'is_scanned': True}
    
    async def extract_from_docx(self, file_path: Path) -> Dict:
        self.logger.info(f"Extracting DOCX: {file_path.name}")
        try:
            doc = DocxDocument(str(file_path))
            full_text = []
            for para in doc.paragraphs:
                text = para.text.strip()
                if text:
                    full_text.append(text)
            return {
                'text': '\n'.join(full_text),
                'pages': [{'page': 1, 'text': '\n'.join(full_text), 'word_count': len(' '.join(full_text).split())}],
                'page_count': 1,
                'is_scanned': False,
                'metadata': {}
            }
        except Exception as e:
            return {'text': f'[Error: {e}]', 'page_count': 0}
    
    async def extract_from_excel(self, file_path: Path) -> Dict:
        self.logger.info(f"Extracting Excel: {file_path.name}")
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            all_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df = pd.DataFrame(ws.values)
                all_text.append(f"Sheet: {sheet_name}")
                all_text.append(df.to_string())
            return {
                'text': '\n'.join(all_text),
                'page_count': len(wb.sheetnames),
                'is_scanned': False,
                'metadata': {}
            }
        except Exception as e:
            return {'text': f'[Error: {e}]', 'page_count': 0}
    
    async def extract_from_csv(self, file_path: Path) -> Dict:
        self.logger.info(f"Extracting CSV: {file_path.name}")
        try:
            df = pd.read_csv(str(file_path))
            return {
                'text': df.to_string(),
                'page_count': 1,
                'is_scanned': False,
                'metadata': {}
            }
        except Exception as e:
            return {'text': f'[Error: {e}]', 'page_count': 0}
